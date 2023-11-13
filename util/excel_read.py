from os import path
import openpyxl

from util.log import _log
from util.cfg_read import cfg

from importlib import import_module
language = getattr(import_module("lang.language", package="lang"), cfg.app_language)

class Devices:
    def __init__(self):
        if path.exists('devices.xlsx'):
            self.devices = self.read()
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(['IP', 'Read Community', 'Write Community'])
            workbook.save('devices.xlsx')

            self.devices = []
            _log._INFO(language.not_find_devices)

    
    def read(self, info=True):
        # 从Excel表格中读取配置
        workbook = openpyxl.load_workbook('devices.xlsx')
        sheet = workbook.active

        devices = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            device = {
                'ip': row[0],
                'read_community': row[1],
                'write_community': row[2]
            }
            # 注释方法
            if row[0][0] == "#":
                continue
            devices.append(device)
        
        if info:
            if devices == []:
                _log._INFO(language.not_find_devices)
            else:
                log_message = f"{language.read_devices_head}"
                for device in devices:
                    log_message += f"\nIP: {device['ip']}, Read Community: {device['read_community']}, Write Community: {device['write_community']}"
                _log._INFO(log_message)

        return devices

    
    def write(self, devices):
        # 将配置写入Excel表格
        workbook = openpyxl.load_workbook('devices.xlsx')
        sheet = workbook.active
        
        for device in devices:
            sheet.append([device['ip'], device['Read Community'], device['Write Community']])
        _log._INFO(f"{language.success_add_device}:\nIP: {device['ip']}, Read Community: {device['Read Community']}, Write Community: {device['Write Community']}")

        workbook.save('devices.xlsx')
        self.devices = self.read(False)

    
    def delete(self, ip):
        # 删除设备
        workbook = openpyxl.load_workbook('devices.xlsx')
        sheet = workbook.active

        delete_row = 0
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == ip:
                # 删除匹配的行
                sheet.delete_rows(row[0].row)
                _log._INFO(f"{language.was_delete}: {ip}")
                delete_row = 1
        if delete_row == 0:
            _log._WARN(language.not_find)

        workbook.save('devices.xlsx')
        self.devices = self.read(False)


devices = Devices()


    
